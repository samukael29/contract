import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import os
from openpyxl import load_workbook

def load_file():
    table = pd.read_excel("Informações.xlsx")
    return table

def dictionary(table, line):
    nome = table.loc[line, "Nome"]
    item1 = table.loc[line, "Item1"]
    item2 = table.loc[line, "Item2"]
    item3 = table.loc[line, "Item3"]

    references = {
        "XXXX": nome,
        "YYYY": item1,
        "ZZZZ": item2,
        "WWWW": item3,
        "DD": str(datetime.now().day),
        "MM": str(datetime.now().month),
        "AAAA": str(datetime.now().year), 
    }
    return references

def create_file(table,header):
    book = openpyxl.Workbook()

    #listas as sheets que ja existem no arquivo
    print(book.sheetnames)

    # criar uma nova sheet
    book.create_sheet('registros2')

    #selecionar uma sheet
    sheet = book['registros2']

    #adicionar dados em linhas
    sheet.append(header)
    for line in table:
        sheet.append(line)

    #salvar arquivo
    book.save('dados.xlsx')


def save_spreadsheet(dataframe):
    writer = pd.ExcelWriter('demo.xlsx', engine='xlsxwriter')
    dataframe.to_excel(writer, sheet_name='Sheet1', index=True)
    writer.save()

def adding_information_to_excel_file1(header,values):
    
    dictionary = {}
    i =0
    for item in header:
        if(str({item})!= str("{'Id'}")):
            dictionary.setdefault(item,values[i])
        i=i+1
    source_dataframe = pd.DataFrame([dictionary])

    print(dictionary)
    dictionary1 = {}
    i =0
    for item in header:
        if(str({item})!= str("{'Id'}")):
            dictionary1.setdefault(item,f"{values[i]}+{values[i]}")
        i=i+1
     
    print(dictionary1)
    new_line_dataframe = pd.DataFrame([dictionary1])
    pd.concat(source_dataframe,new_line_dataframe)
    print(dataframe)

    # save_spreadsheet(dataframe)

def adding_information_to_excel_file(header,values):
    
    workbook_name = 'sample.xlsx'
    id = 0
    if (os.path.exists(workbook_name)):
        wb = load_workbook(workbook_name)
        page = wb.active
        id = (page.max_row)
    else:
        wb = openpyxl.Workbook()
        page = wb.active
        page.title = 'companies'
        page.append(header)
        id = 1
    
    values[0] = id
    page.append(values)

    wb.save(filename=workbook_name)

def testfile():
    workbook_name = 'sample.xlsx'
    id = 0
    if (os.path.exists(workbook_name)):
        wb = load_workbook(workbook_name)
        page = wb.active
        id = (page.max_row) - 1

# testfile()